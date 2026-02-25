import pandas as pd; 

file = pd.read_json('../Documents/embedded_data.json')
file = file.rename(columns={'embedded' : 'Embed Status'})
file = file.sort_values(by='Embed Status').reset_index(drop=True)
#file
# ?print(file)

excel_path = '../Documents/Embedded_Dataset.xlsx'
#exporting to excel
file.to_excel(excel_path, index=False, engine='openpyxl')
print("Updated the Excel File!")

#! Beautifying the Dataset
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


wb = load_workbook('./Documents/Embedded_Dataset.xlsx')
sheet = wb.active

#? Setting widths of the columns
sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 20

sheet.row_dimensions[1].height = 50
#?setting alignment of the contents in the columns
align_2 = Alignment(horizontal = 'center' , vertical= 'center' , wrapText=True)


for cell in sheet['A']:
    cell.alignment = align_2

for cell in sheet['B']:
    cell.alignment = align_2

for cell in sheet['C']:
    cell.alignment = align_2

for cell in sheet['D']:
    cell.alignment = align_2

#?formatting the header row and the other rows
for cell in sheet[1][:4]: 
    cell.font = Font(size=14, bold=True)

for row_num in range(2 , sheet.max_row + 1):
    sheet.row_dimensions[row_num].height = 30

#applying the changes
wb.save('./Documents/Embedded_Dataset.xlsx')
print("The styling is also done!")